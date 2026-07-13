"""
Legacy-bot user import — reads a CSV or Excel file of existing Telegram
users (chat_id, username, full_name, balance, created_at) and merges them
into the local `users` table, keyed on telegram_id so the bot can keep
messaging/notifying the exact same Chat IDs after taking over an old bot.

Three admin-selectable modes control how existing rows are handled:
  - "new_only"       — only create users that don't exist yet; never touch
                        an existing row.
  - "update_info"    — create new users; for existing users, refresh
                        username/full_name only (balance is left untouched).
  - "update_balance" — same as update_info, but also overwrites balance for
                        existing users.

New users are always created with the imported balance (there is nothing to
"overwrite" yet). Rows are never deleted; nothing outside the touched rows
is affected.
"""
import csv
import io
import logging
from datetime import datetime
from typing import Optional

from sqlalchemy.orm import Session
from models import User

logger = logging.getLogger(__name__)

VALID_MODES = {"new_only", "update_info", "update_balance"}

_DATE_FORMATS = [
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d",
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y",
    "%m/%d/%Y %H:%M:%S",
    "%m/%d/%Y",
]

# Accepted header aliases (case-insensitive) -> canonical field name.
_HEADER_ALIASES = {
    "chat_id": "chat_id",
    "chatid": "chat_id",
    "telegram_chat_id": "chat_id",
    "telegram_id": "chat_id",
    "id": "chat_id",
    "username": "username",
    "user_name": "username",
    "full_name": "full_name",
    "fullname": "full_name",
    "name": "full_name",
    "balance": "balance",
    "created_at": "created_at",
    "created": "created_at",
    "joined_at": "created_at",
}


def _parse_date(raw) -> Optional[datetime]:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    raw = str(raw).strip()
    if not raw:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def _parse_balance(raw) -> float:
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    raw = str(raw).strip().replace(",", "").replace(" ", "")
    if not raw:
        return 0.0
    try:
        return float(raw)
    except ValueError:
        return 0.0


def _split_full_name(full_name: str) -> tuple[str, str]:
    """Users.first_name/last_name is what the schema has; split a single
    "full_name" import column into those two on the first space."""
    full_name = (full_name or "").strip()
    if not full_name:
        return "", ""
    parts = full_name.split(" ", 1)
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[1]


def parse_import_file(filename: str, content: bytes) -> list[dict]:
    """Returns a list of raw row dicts with canonical keys
    (chat_id/username/full_name/balance/created_at). Raises ValueError on an
    unsupported file type or unreadable content."""
    lower = (filename or "").lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return _parse_excel(content)
    if lower.endswith(".csv") or lower.endswith(".txt"):
        return _parse_csv(content)
    # Fall back to sniffing: try Excel magic bytes, else assume CSV text.
    if content[:2] == b"PK":
        return _parse_excel(content)
    return _parse_csv(content)


def _normalize_headers(raw_headers: list) -> dict:
    """Maps column index -> canonical field name for recognized headers."""
    mapping = {}
    for idx, h in enumerate(raw_headers):
        key = str(h or "").strip().lower().replace(" ", "_")
        canonical = _HEADER_ALIASES.get(key)
        if canonical:
            mapping[idx] = canonical
    return mapping


def _parse_csv(content: bytes) -> list[dict]:
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1258", "latin-1"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Không thể đọc file CSV (encoding không hỗ trợ)")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = list(reader)
    if not rows:
        return []
    header_map = _normalize_headers(rows[0])
    if not header_map:
        raise ValueError("Không tìm thấy cột hợp lệ (chat_id, username, full_name, balance, created_at)")
    out = []
    for row in rows[1:]:
        if not any(c.strip() for c in row if isinstance(c, str)):
            continue
        rec = {}
        for idx, canonical in header_map.items():
            if idx < len(row):
                rec[canonical] = row[idx]
        out.append(rec)
    return out


def _parse_excel(content: bytes) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    header_map = _normalize_headers(list(header_row))
    if not header_map:
        raise ValueError("Không tìm thấy cột hợp lệ (chat_id, username, full_name, balance, created_at)")
    out = []
    for row in rows_iter:
        if row is None or all(c is None for c in row):
            continue
        rec = {}
        for idx, canonical in header_map.items():
            if idx < len(row):
                rec[canonical] = row[idx]
        out.append(rec)
    return out


def import_users(db: Session, rows: list[dict], mode: str) -> dict:
    """Applies the parsed rows to the users table per `mode`. Returns a
    stats dict: total, created, updated, duplicates, errors, error_rows
    (first few, for admin review), total_balance_imported."""
    if mode not in VALID_MODES:
        raise ValueError(f"Invalid import mode: {mode}")

    total = len(rows)
    created = 0
    updated = 0
    duplicates = 0
    errors = 0
    error_rows: list[str] = []
    total_balance_imported = 0.0

    for i, row in enumerate(rows, start=1):
        raw_chat_id = row.get("chat_id")
        chat_id = str(raw_chat_id).strip() if raw_chat_id is not None else ""
        # Excel may hand back chat ids as floats (e.g. 123456.0) — normalize.
        if chat_id.endswith(".0") and chat_id.replace(".0", "").lstrip("-").isdigit():
            chat_id = chat_id[:-2]
        if not chat_id or not chat_id.lstrip("-").isdigit():
            errors += 1
            if len(error_rows) < 20:
                error_rows.append(f"Dòng {i}: thiếu hoặc sai chat_id ({raw_chat_id!r})")
            continue

        username = (row.get("username") or "").strip() or None
        if username and username.startswith("@"):
            username = username[1:]
        full_name = (row.get("full_name") or "").strip()
        first_name, last_name = _split_full_name(full_name)
        balance = _parse_balance(row.get("balance"))
        created_at = _parse_date(row.get("created_at"))

        try:
            existing = db.query(User).filter(User.telegram_id == chat_id).first()
            if existing:
                duplicates += 1
                if mode == "new_only":
                    continue
                if username:
                    existing.username = username
                if full_name:
                    existing.first_name = first_name
                    existing.last_name = last_name
                if mode == "update_balance":
                    existing.balance = balance
                    total_balance_imported += balance
                updated += 1
            else:
                user = User(
                    telegram_id=chat_id,
                    username=username,
                    first_name=first_name or None,
                    last_name=last_name or None,
                    balance=balance,
                )
                if created_at:
                    user.created_at = created_at
                db.add(user)
                created += 1
                total_balance_imported += balance
        except Exception as e:
            db.rollback()
            errors += 1
            if len(error_rows) < 20:
                error_rows.append(f"Dòng {i}: {e}")
            continue

    db.commit()
    logger.info(
        f"USER_IMPORT: mode={mode} total={total} created={created} updated={updated} "
        f"duplicates={duplicates} errors={errors} balance_imported={total_balance_imported}"
    )
    return {
        "total": total,
        "created": created,
        "updated": updated,
        "success": created + updated,
        "duplicates": duplicates,
        "errors": errors,
        "error_rows": error_rows,
        "total_balance_imported": total_balance_imported,
        "mode": mode,
    }
